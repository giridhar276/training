import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Step 1: Connect to employee_db database
connection = pymysql.connect(
    host="localhost",
    user="root",
    password="your_mysql_password",
    database="employee_db"
)

cursor = connection.cursor()

# Step 2: Read data from MySQL table
select_query = """
SELECT
    id,
    age,
    workclass,
    fnlwgt,
    education,
    educational_num,
    marital_status,
    occupation,
    relationship,
    race,
    gender,
    capital_gain,
    capital_loss,
    hours_per_week,
    native_country,
    income
FROM emp_info
"""

cursor.execute(select_query)
rows = cursor.fetchall()

# Step 3: Column names for Excel
columns = [
    "id",
    "age",
    "workclass",
    "fnlwgt",
    "education",
    "educational_num",
    "marital_status",
    "occupation",
    "relationship",
    "race",
    "gender",
    "capital_gain",
    "capital_loss",
    "hours_per_week",
    "native_country",
    "income"
]

# Step 4: Create Excel workbook
workbook = Workbook()
sheet = workbook.active
sheet.title = "Employee Data"

# Step 5: Write header row
sheet.append(columns)

# Step 6: Write data rows
for row in rows:
    sheet.append(row)

# Step 7: Basic Excel formatting
header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
header_font = Font(bold=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for cell in sheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

for row in sheet.iter_rows(min_row=2):
    for cell in row:
        cell.border = thin_border

# Auto-adjust column width
for column_cells in sheet.columns:
    max_length = 0
    column_letter = get_column_letter(column_cells[0].column)

    for cell in column_cells:
        if cell.value is not None:
            max_length = max(max_length, len(str(cell.value)))

    sheet.column_dimensions[column_letter].width = max_length + 2

# Step 8: Save Excel file
workbook.save("employee_output.xlsx")

print("Data exported successfully to employee_output.xlsx")

# Step 9: Close resources
cursor.close()
connection.close()
